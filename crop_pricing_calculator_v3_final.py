from __future__ import annotations

# ════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION  —  only edit this section
# ════════════════════════════════════════════════════════════════════════════

INPUT_PATH = r"D:\Crop Insurance Pricing\input\Karnataka Kharif Pricing.xlsx"
OUTPUT_DIR = r"D:\Crop Insurance Pricing\input"
SI_PATH    = r"D:\Crop Insurance Pricing\input\Suminsured.xlsx"

# ════════════════════════════════════════════════════════════════════════════
#  IMPORTS & CONSTANTS
# ════════════════════════════════════════════════════════════════════════════

import datetime, os, re
import numpy as np
import pandas as pd
from scipy import stats
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional

DETREND_BASE_YEAR    = 2024
CL_WINDOW            = 7
SLOPE_WINDOW         = 10
BURN_WINDOWS         = [3, 5, 10, 11, 12, 13, 14]
BC_WINDOWS           = [10, 11, 12, 13, 14]
CORRIDOR_YEAR_WINDOW = 10
CAP_ONLY_CEILING     = 1.10   # cap-only always uses 110% ceiling (verified against Excel)
CORRIDOR_MODELS      = [(80, 110), (60, 130)]
LOSS_GROUP_KEYS      = ['Cluster', 'Season', 'State', 'District', 'Crop Name as per BOQ']
SI_JOIN_KEYS         = ['Cluster', 'Season', 'State', 'District', 'Crop Name as per BOQ']

# ════════════════════════════════════════════════════════════════════════════
#  SHARED HELPERS
# ════════════════════════════════════════════════════════════════════════════

def _safe_div(a, b):
    return float(a / b) if (pd.notna(a) and pd.notna(b) and b != 0) else np.nan

def _safe_mean(lst):
    v = [x for x in lst if pd.notna(x)]
    return float(np.mean(v)) if v else np.nan

def _nan_none(x):
    if x is None: return None
    try:
        if np.isnan(x): return None
    except (TypeError, ValueError): pass
    return x

def _cluster_display(val):
    """'CL N' -> 'Cluster N' for Excel calc-block labels."""
    m = re.match(r'CL\s+(\d+)', str(val))
    return f'Cluster {m.group(1)}' if m else val

def detect_year_columns(df: pd.DataFrame) -> List[int]:
    years = []
    for col in df.columns:
        try:
            y = int(col)
            if 1900 < y <= DETREND_BASE_YEAR: years.append(y)
        except (ValueError, TypeError): pass
    return sorted(years)

def detect_id_columns(df: pd.DataFrame, year_ints: List[int]) -> List[str]:
    year_strs = {str(y) for y in year_ints} | set(year_ints)
    id_cols = []
    for col in df.columns:
        if col in year_strs: break
        if col != 'IL': id_cols.append(col)
    return id_cols

# ════════════════════════════════════════════════════════════════════════════
#  BASE DATA ROW-LEVEL HELPERS
# ════════════════════════════════════════════════════════════════════════════

def _last_data_year(row_yields, all_years):
    for y in reversed(all_years):
        v = row_yields.get(y)
        if v is not None and not np.isnan(v): return y
    return None

def _window_years(last_year, n, all_years):
    start = last_year - n + 1
    return [y for y in all_years if start <= y <= last_year]

def _yields_arr(row_yields, year_list):
    return np.array([row_yields.get(y, np.nan) for y in year_list], dtype=float)

def _valid(arr):   return arr[~np.isnan(arr)]
def _nsmallest(arr, n): v = np.sort(_valid(arr)); return float(v[n-1]) if len(v) >= n else np.nan
def _nlargest(arr, n):  v = np.sort(_valid(arr))[::-1]; return float(v[n-1]) if len(v) >= n else np.nan

def _cl_window(all_years): return all_years[-CL_WINDOW:]

def _cl_yields(row_yields, all_years):
    arr = _yields_arr(row_yields, _cl_window(all_years))
    return _nsmallest(arr, 1), _nsmallest(arr, 2)

def _cl_years(row_yields, cl1_val, cl2_val, all_years):
    win = _cl_window(all_years)
    def nth_year(target, occ):
        hits = [y for y in win if row_yields.get(y) == target]
        return hits[occ-1] if len(hits) >= occ else 0
    if np.isnan(cl1_val): return 0, 0
    cl1_y = nth_year(cl1_val, 1)
    if np.isnan(cl2_val): return cl1_y, 0
    return cl1_y, (nth_year(cl2_val, 2) if cl1_val == cl2_val else nth_year(cl2_val, 1))

def _average_rev(row_yields, all_years):
    arr  = _yields_arr(row_yields, _cl_window(all_years))
    top5 = [v for i in range(1, 6) if not np.isnan(v := _nlargest(arr, i))]
    return float(np.mean(top5)) if top5 else np.nan

def _claim(yield_val, ty):
    if np.isnan(yield_val) or np.isnan(ty) or ty == 0: return np.nan
    return max(0.0, (ty - yield_val) / ty)

def _trend_stats(row_yields, last_year, all_years):
    win   = _window_years(last_year, SLOPE_WINDOW, all_years)
    x_all = np.array(win, dtype=float); y_all = _yields_arr(row_yields, win)
    mask  = ~np.isnan(y_all); x, y = x_all[mask], y_all[mask]
    empty = {k: np.nan for k in ('SLOPE','RSQ','STD DEV','CV','T VALUE','P VALUE')}
    empty.update({'Count RSQ': 0, 'SIGN TREND?': False})
    if len(y) < 2: return empty
    slope, _, r, _, se = stats.linregress(x, y)
    rsq   = r**2; std_dev = float(np.std(y, ddof=1)); mean_y = float(np.mean(y))
    cv    = std_dev / mean_y if mean_y != 0 else np.nan
    t_val = slope / se if se != 0 else np.nan
    p_val = float(stats.t.sf(abs(t_val), df=len(y))*2) if not np.isnan(t_val) else np.nan
    return {'SLOPE': float(slope), 'RSQ': float(rsq), 'Count RSQ': 1 if rsq > 0.7 else 0,
            'STD DEV': std_dev, 'CV': cv, 'T VALUE': float(t_val), 'P VALUE': p_val,
            'SIGN TREND?': bool(p_val < 0.05) if not np.isnan(p_val) else False}

def _detrend(yield_val, year, slope, p_value, cl1_year, cl2_year):
    if np.isnan(yield_val): return ''
    if np.isnan(p_value) or p_value > 0.05: return yield_val
    if year in (cl1_year, cl2_year): return yield_val
    return max(0.0, yield_val + slope * (DETREND_BASE_YEAR - year))

def _burn_cost(claims, last_year, all_years, n):
    win  = _window_years(last_year, n, all_years)
    vals = [claims[y] for y in win if y in claims and not np.isnan(claims[y])]
    return float(np.mean(vals)) if vals else np.nan

def _calc_row(row_yields, il, all_years):
    out       = {}
    last_year = _last_data_year(row_yields, all_years)
    if last_year is None: return out
    cl1_yield, cl2_yield = _cl_yields(row_yields, all_years)
    cl1_year,  cl2_year  = _cl_years(row_yields, cl1_yield, cl2_yield, all_years)
    out.update({'CL 1 Yield': cl1_yield, 'CL 2 Yield': cl2_yield,
                'CL 1': cl1_year, 'CL 2': cl2_year})
    avg = _average_rev(row_yields, all_years)
    ty  = il * avg if (not np.isnan(il) and not np.isnan(avg)) else np.nan
    out.update({'Average (Rev Guidelines)': avg, 'IL': il, 'TY (Rev Guidelines)': ty})
    claims_normal = {}
    for y in all_years:
        c = _claim(row_yields.get(y, np.nan), ty)
        out[f'{y} Claims'] = c; claims_normal[y] = c
    for n in BURN_WINDOWS:
        out[f'Burn Cost Normal @ {n} yrs'] = _burn_cost(claims_normal, last_year, all_years, n)
    ts = _trend_stats(row_yields, last_year, all_years)
    out.update(ts)
    detrend_yields = {}
    for y in all_years:
        dt = _detrend(row_yields.get(y, np.nan), y, ts['SLOPE'], ts['P VALUE'], cl1_year, cl2_year)
        out[f'{y} DeTrend @ T TEST'] = dt
        detrend_yields[y] = dt if dt != '' else np.nan
    claims_ttest = {}
    for y in all_years:
        dt = detrend_yields.get(y, np.nan)
        c  = _claim(dt, ty) if not np.isnan(dt) else np.nan
        out[f'{y} DeTrend @ T TEST Claims'] = c; claims_ttest[y] = c
    for n in BURN_WINDOWS:
        out[f'Burn Cost @ T TEST @ {n} yrs'] = _burn_cost(claims_ttest, last_year, all_years, n)
    return out

# ════════════════════════════════════════════════════════════════════════════
#  SHEET 1: Base Data & T Test
# ════════════════════════════════════════════════════════════════════════════

def calculate_all(df: pd.DataFrame) -> pd.DataFrame:
    df        = df.copy()
    year_ints = detect_year_columns(df)
    id_cols   = detect_id_columns(df, year_ints)
    if not year_ints: raise ValueError("No year columns detected.")
    if 'IL' not in df.columns: raise ValueError("'IL' column not found.")
    df.rename(columns={y: str(y) for y in year_ints if y in df.columns}, inplace=True)
    for y in year_ints: df[str(y)] = pd.to_numeric(df[str(y)], errors='coerce')
    output_rows = []
    for rec in df.to_dict('records'):
        row_yields = {}
        for y in year_ints:
            v = rec.get(str(y))
            try: row_yields[y] = float(v) if v is not None else np.nan
            except (TypeError, ValueError): row_yields[y] = np.nan
        try: il = float(rec.get('IL')) if rec.get('IL') is not None else np.nan
        except (TypeError, ValueError): il = np.nan
        output_rows.append({**rec, **_calc_row(row_yields, il, year_ints)})
    result  = pd.DataFrame(output_rows)
    years   = [str(y) for y in year_ints]
    ordered = (id_cols + years
        + ['CL 1 Yield','CL 2 Yield','CL 1','CL 2',
           'Average (Rev Guidelines)','IL','TY (Rev Guidelines)']
        + [f'{y} Claims' for y in years]
        + [f'Burn Cost Normal @ {n} yrs' for n in BURN_WINDOWS]
        + ['SLOPE','RSQ','Count RSQ','STD DEV','CV','T VALUE','P VALUE','SIGN TREND?']
        + [f'{y} DeTrend @ T TEST' for y in years]
        + [f'{y} DeTrend @ T TEST Claims' for y in years]
        + [f'Burn Cost @ T TEST @ {n} yrs' for n in BURN_WINDOWS])
    present    = [c for c in ordered if c in result.columns]
    known      = set(ordered)
    user_extra = [c for c in result.columns if c not in known]
    if id_cols and id_cols[-1] in present:
        idx     = present.index(id_cols[-1]) + 1
        present = present[:idx] + user_extra + present[idx:]
    else:
        present = user_extra + present
    return result[present]

# ════════════════════════════════════════════════════════════════════════════
#  SHEET 2: LossCost
# ════════════════════════════════════════════════════════════════════════════

def calculate_losscost(df_base: pd.DataFrame) -> pd.DataFrame:
    year_ints = detect_year_columns(df_base); years = [str(y) for y in year_ints]
    all_agg   = ([f'{y} Claims' for y in years]
               + [f'Burn Cost Normal @ {n} yrs' for n in BURN_WINDOWS]
               + [f'{y} DeTrend @ T TEST Claims' for y in years]
               + [f'Burn Cost @ T TEST @ {n} yrs' for n in BURN_WINDOWS])
    missing = [k for k in LOSS_GROUP_KEYS if k not in df_base.columns]
    if missing: raise ValueError(f"LossCost grouping keys not found: {missing}")
    df_work  = df_base[LOSS_GROUP_KEYS + [c for c in all_agg if c in df_base.columns]].copy()
    agg_pres = [c for c in all_agg if c in df_work.columns]
    for col in agg_pres: df_work[col] = pd.to_numeric(df_work[col], errors='coerce')
    df_g = df_work.groupby(LOSS_GROUP_KEYS, sort=False)[agg_pres].mean().reset_index()
    rename = ({f'{y} Claims': f'Average of {y} Claims' for y in years}
            | {f'Burn Cost Normal @ {n} yrs': f'Average of Burn Cost Normal @ {n} yrs' for n in BURN_WINDOWS}
            | {f'{y} DeTrend @ T TEST Claims': f'Average of {y} DeTrend @ T TEST Claims' for y in years}
            | {f'Burn Cost @ T TEST @ {n} yrs': f'Average of Burn Cost @ T TEST @ {n} yrs' for n in BURN_WINDOWS})
    df_g.rename(columns=rename, inplace=True)
    ordered = (LOSS_GROUP_KEYS
        + [f'Average of {y} Claims' for y in years]
        + [f'Average of Burn Cost Normal @ {n} yrs' for n in BURN_WINDOWS]
        + [f'Average of {y} DeTrend @ T TEST Claims' for y in years]
        + [f'Average of Burn Cost @ T TEST @ {n} yrs' for n in BURN_WINDOWS])
    return df_g[[c for c in ordered if c in df_g.columns]]

# ════════════════════════════════════════════════════════════════════════════
#  SHEET 3: BC
# ════════════════════════════════════════════════════════════════════════════

def calculate_bc(df_losscost: pd.DataFrame) -> pd.DataFrame:
    keep = (LOSS_GROUP_KEYS
          + [f'Average of Burn Cost Normal @ {n} yrs' for n in BC_WINDOWS]
          + [f'Average of Burn Cost @ T TEST @ {n} yrs' for n in BC_WINDOWS])
    return df_losscost[[c for c in keep if c in df_losscost.columns]].copy().reset_index(drop=True)

# ════════════════════════════════════════════════════════════════════════════
#  SHEET 4: Suminsured
# ════════════════════════════════════════════════════════════════════════════

def load_suminsured(si_path: str) -> pd.DataFrame:
    df_si = pd.read_csv(si_path) if si_path.lower().endswith('.csv') else pd.read_excel(si_path)
    missing = [k for k in SI_JOIN_KEYS if k not in df_si.columns]
    if missing: raise ValueError(f"Suminsured missing join keys: {missing}")
    excluded    = set(SI_JOIN_KEYS) | {'Concat'}
    si_val_cols = [c for c in df_si.columns
                   if c not in excluded
                   and pd.api.types.is_numeric_dtype(df_si[c])
                   and df_si[c].notna().any()]
    if not si_val_cols: raise ValueError("No SI valuation columns found.")
    print(f"  Suminsured SI columns: {si_val_cols}")
    return df_si[[c for c in SI_JOIN_KEYS + si_val_cols if c in df_si.columns]].copy()

# ════════════════════════════════════════════════════════════════════════════
#  SHEET 5: CupnCapModel
# ════════════════════════════════════════════════════════════════════════════

def calculate_cupncap(df_losscost: pd.DataFrame, df_si: pd.DataFrame, si_col: str) -> pd.DataFrame:
    if si_col not in df_si.columns: raise ValueError(f"SI column '{si_col}' not found.")
    year_pat  = re.compile(r'^Average of (\d{4}) Claims$')
    ttest_pat = re.compile(r'^Average of (\d{4}) DeTrend @ T TEST Claims$')
    normal_year_strs = sorted([m.group(1) for c in df_losscost.columns if (m := year_pat.match(c))])
    ttest_year_strs  = sorted([m.group(1) for c in df_losscost.columns if (m := ttest_pat.match(c))])
    df = df_losscost.merge(df_si[SI_JOIN_KEYS + [si_col]], on=SI_JOIN_KEYS, how='left')
    normal_claim_cols = [f'Average of {y} Claims' for y in normal_year_strs]
    ttest_claim_cols  = [f'Average of {y} DeTrend @ T TEST Claims' for y in ttest_year_strs]
    nb_lbl = [f'Normal BC Year {i}' for i in range(1, CORRIDOR_YEAR_WINDOW+1)]
    tt_lbl = [f'T-TestBC Year {i}'  for i in range(1, CORRIDOR_YEAR_WINDOW+1)]
    output_rows = []
    for _, row in df.iterrows():
        si_val = row.get(si_col, np.nan)
        def pick_window(claim_cols):
            sel = []
            for col in reversed(claim_cols):
                if pd.notna(row.get(col, np.nan)): sel.append(row.get(col))
                if len(sel) == CORRIDOR_YEAR_WINDOW: break
            return list(reversed(sel))
        def to_premium(vals):
            if pd.isna(si_val): return [np.nan]*len(vals)
            return [v*si_val if pd.notna(v) else np.nan for v in vals]
        norm_prm  = to_premium(pick_window(normal_claim_cols))
        ttest_prm = to_premium(pick_window(ttest_claim_cols))
        rec = {k: row.get(k) for k in LOSS_GROUP_KEYS}
        rec[si_col] = si_val
        for i, lbl in enumerate(nb_lbl): rec[lbl] = norm_prm[i] if i < len(norm_prm) else np.nan
        rec['Average 10 Year  BC'] = _safe_mean(norm_prm)
        for i, lbl in enumerate(tt_lbl): rec[lbl] = ttest_prm[i] if i < len(ttest_prm) else np.nan
        rec['Average 10 Year  BC T-Test'] = _safe_mean(ttest_prm)
        output_rows.append(rec)
    df_out  = pd.DataFrame(output_rows)
    ordered = (LOSS_GROUP_KEYS + [si_col] + nb_lbl
               + ['Average 10 Year  BC'] + tt_lbl + ['Average 10 Year  BC T-Test'])
    return df_out[[c for c in ordered if c in df_out.columns]].reset_index(drop=True)

# ════════════════════════════════════════════════════════════════════════════
#  SHEET 6+: Corridor Model (80-110 and 60-130)
#
#  Each sheet has 3 stacked sections: Overall, Kharif, Rabi
#    • Overall  — no season filter, all rows aggregated by Cluster
#    • Kharif   — only Season == 'Kharif' rows, own pivot + own calc block
#    • Rabi     — only Season == 'Rabi'   rows, own pivot + own calc block
#  Sections with no data are skipped automatically.
#  Cap-only always uses CAP_ONLY_CEILING (110%) regardless of corridor model.
#  85-column layout verified cell-by-cell against reference Excel.
# ════════════════════════════════════════════════════════════════════════════

def _agg_by_cluster(df_cupncap: pd.DataFrame, si_col: str,
                    season_filter: Optional[str] = None) -> pd.DataFrame:
    """
    Sum CupnCapModel rows by Cluster, optionally filtered to one season.
    season_filter=None -> Overall (all seasons combined).
    Returns a DataFrame indexed by Cluster with summed numeric columns.
    """
    nb_lbl = [f'Normal BC Year {i}' for i in range(1, CORRIDOR_YEAR_WINDOW+1)]
    tt_lbl = [f'T-TestBC Year {i}'  for i in range(1, CORRIDOR_YEAR_WINDOW+1)]
    num_cols = [si_col] + nb_lbl + ['Average 10 Year  BC'] + tt_lbl + ['Average 10 Year  BC T-Test']

    df = df_cupncap.copy()
    if season_filter is not None:
        df = df[df['Season'].str.strip().str.lower() == season_filter.lower()]
    if df.empty: return pd.DataFrame()

    for c in num_cols:
        if c in df.columns: df[c] = pd.to_numeric(df[c], errors='coerce')
    return df.groupby('Cluster')[[c for c in num_cols if c in df.columns]].sum(min_count=1).reset_index()


def _cluster_metrics(pivot_row: pd.Series, si_col: str, floor: float, cap: float) -> Dict:
    """Compute all corridor metrics for one cluster row from the pivoted DataFrame."""
    N      = CORRIDOR_YEAR_WINDOW
    nb_lbl = [f'Normal BC Year {i}' for i in range(1, N+1)]
    tt_lbl = [f'T-TestBC Year {i}'  for i in range(1, N+1)]
    esi    = pivot_row.get(si_col, np.nan)
    avg_n  = pivot_row.get('Average 10 Year  BC', np.nan)
    avg_t  = pivot_row.get('Average 10 Year  BC T-Test', np.nan)
    bc_n   = _safe_div(avg_n, esi)
    bc_t   = _safe_div(avg_t, esi)

    def per_year(lbl_list, avg_ref, bc_rate):
        yr_sums, bc_yrs, rev_lr, capped, caponly = [], [], [], [], []
        for lbl in lbl_list:
            sy = pivot_row.get(lbl, np.nan)
            yr_sums.append(sy)
            bc_yrs.append(_safe_div(sy, esi))
            rl = _safe_div(sy, avg_ref)
            rev_lr.append(rl)
            if pd.notna(rl) and pd.notna(bc_rate):
                capped.append(max(floor, min(cap, rl)) * bc_rate)
                caponly.append(min(CAP_ONLY_CEILING, rl) * bc_rate)
            else:
                capped.append(np.nan); caponly.append(np.nan)
        return yr_sums, bc_yrs, rev_lr, capped, caponly

    n_sums, n_bc, n_rl, nc, nco = per_year(nb_lbl, avg_n, bc_n)
    t_sums, t_bc, t_rl, tc, tco = per_year(tt_lbl, avg_t, bc_t)
    return dict(
        sum_esi=esi, sum_avg10_norm=avg_n, sum_avg10_ttest=avg_t,
        norm_yr_sums=n_sums, tt_yr_sums=t_sums,
        bc_normal=bc_n, bc_ttest=bc_t,
        norm_bc_yrs=n_bc, norm_rev_lr=n_rl,
        tt_bc_yrs=t_bc,   tt_rev_lr=t_rl,
        norm_capped=nc,   norm_caponly=nco,
        tt_capped=tc,     tt_caponly=tco,
        corr_avg_norm=_safe_mean(nc),  caponly_avg_norm=_safe_mean(nco),
        corr_avg_tt=_safe_mean(tc),    caponly_avg_tt=_safe_mean(tco),
    )


def _write_section(ws, start_row: int, clusters_data: Dict,
                   floor_pct: int, cap_pct: int,
                   section_label: Optional[str] = None) -> int:
    """
    Write one self-contained corridor section (85 columns) into ws.
    Returns the next available row after a 2-row gap.

    Column layout (1-based, verified against reference Excel):
      1       Row Labels / cluster pivot label
      2-11    Sum of Normal BC Year 1-10  [Pivot 1 — this section's data]
      12      Sum of Average 10 Year BC
      13      Sum of ESI
      15      Cluster display label       [Pivot 2 — same data, different label style]
      16-25   Sum of Normal BC Year 1-10
      26-27   Sum Avg 10yr BC / Sum ESI
      29      Cluster  [Calc block]
      30-38   BC Year 1-9
      39      10 Year BC  (= BC Year 10)
      40      BC Normal @ 10 yrs
      42-51   Revised LR Year 1-10
      54      Cluster  [Capped BC block]
      55-64   BC Year 1-10 (capped: floor-cap corridor)
      65      Capped avg label
      68      Cluster  [Cap-only BC block — always 110% ceiling]
      69-78   BC Year 1-10 (cap-only)
      79      Cap-only avg label ("110% Avg. Normal BC")
      82      Cluster  [Summary]
      83      BC Normal @ 10 yrs
      84      Corridor Avg Normal BC
      85      Cap-only Avg (label uses cap_pct, value always 110%)
    """
    N       = CORRIDOR_YEAR_WINDOW
    cl      = f'{floor_pct}%-{cap_pct}%'
    clusters = sorted(clusters_data.keys())
    r0, r1, r2, r3 = start_row, start_row+1, start_row+2, start_row+3
    r_total = r3 + len(clusters)

    # Section label row (blank for Overall, 'Season / Kharif' for sub-seasons)
    if section_label:
        ws.cell(row=r0, column=1).value = 'Season'
        ws.cell(row=r0, column=2).value = section_label

    ws.cell(row=r1, column=54).value = cl
    ws.cell(row=r1, column=70).value = cap_pct / 100

    # ── Headers ─────────────────────────────────────────────────────────────
    ws.cell(row=r2, column=1).value = 'Row Labels'
    for i in range(N):
        ws.cell(row=r2, column=2+i).value  = f'Sum of Normal BC Year {i+1}'
        ws.cell(row=r2, column=16+i).value = f'Sum of Normal BC Year {i+1}'
        ws.cell(row=r2, column=55+i).value = f'BC Year {i+1}'
        ws.cell(row=r2, column=69+i).value = f'BC Year {i+1}'
    for i in range(N-1):
        ws.cell(row=r2, column=30+i).value = f'BC Year {i+1}'
        ws.cell(row=r2, column=42+i).value = f'Revised LR  Year {i+1}'
    for col, val in [(12,'Sum of Average 10 Year  BC'), (13,'Sum of ESI'),
                     (26,'Sum of Average 10 Year  BC'), (27,'Sum of ESI'),
                     (29,'Cluster'), (39,'10 Year BC'),  (40,'BC Normal @ 10 yrs'),
                     (51,'Revised LR Year 10'),          (54,'Cluster'),
                     (68,'Cluster'),                     (82,'Cluster'),
                     (83,'BC Normal @ 10 yrs')]:
        ws.cell(row=r2, column=col).value = val
    ws.cell(row=r2, column=65).value = (f'{cl} Avg. Normal BC' if floor_pct == 80
                                         else f'{cl} 10 Avg. Normal BC')
    ws.cell(row=r2, column=79).value = ('110% Avg. Normal BC' if floor_pct == 80
                                         else '110% 10 Avg. Normal BC')
    ws.cell(row=r2, column=84).value = f'{cl} Avg. Normal BC'
    ws.cell(row=r2, column=85).value = f'{cap_pct}% Avg. Normal BC'

    # ── Data rows ───────────────────────────────────────────────────────────
    g_yr  = [0.0]*N; g_avg = 0.0; g_esi = 0.0
    for idx, cluster in enumerate(clusters):
        d  = clusters_data[cluster]
        r  = r3 + idx
        cd = _cluster_display(cluster)

        # Pivot 1 — raw cluster label in col A
        ws.cell(row=r, column=1).value  = cluster
        for i in range(N): ws.cell(row=r, column=2+i).value  = _nan_none(d['norm_yr_sums'][i])
        ws.cell(row=r, column=12).value = _nan_none(d['sum_avg10_norm'])
        ws.cell(row=r, column=13).value = _nan_none(d['sum_esi'])

        # Pivot 2 — 'Cluster N' display label in col 15
        ws.cell(row=r, column=15).value = cd
        for i in range(N): ws.cell(row=r, column=16+i).value = _nan_none(d['norm_yr_sums'][i])
        ws.cell(row=r, column=26).value = _nan_none(d['sum_avg10_norm'])
        ws.cell(row=r, column=27).value = _nan_none(d['sum_esi'])

        # Calc block
        ws.cell(row=r, column=29).value = cd
        for i in range(N-1): ws.cell(row=r, column=30+i).value = _nan_none(d['norm_bc_yrs'][i])
        ws.cell(row=r, column=39).value = _nan_none(d['norm_bc_yrs'][N-1])
        ws.cell(row=r, column=40).value = _nan_none(d['bc_normal'])

        # Revised LR
        for i in range(N): ws.cell(row=r, column=42+i).value = _nan_none(d['norm_rev_lr'][i])

        # Capped BC
        ws.cell(row=r, column=54).value = cd
        for i in range(N): ws.cell(row=r, column=55+i).value = _nan_none(d['norm_capped'][i])
        ws.cell(row=r, column=65).value = _nan_none(d['corr_avg_norm'])

        # Cap-only BC
        ws.cell(row=r, column=68).value = cd
        for i in range(N): ws.cell(row=r, column=69+i).value = _nan_none(d['norm_caponly'][i])
        ws.cell(row=r, column=79).value = _nan_none(d['caponly_avg_norm'])

        # Summary
        ws.cell(row=r, column=82).value = cd
        ws.cell(row=r, column=83).value = _nan_none(d['bc_normal'])
        ws.cell(row=r, column=84).value = _nan_none(d['corr_avg_norm'])
        ws.cell(row=r, column=85).value = _nan_none(d['caponly_avg_norm'])

        # Accumulators for grand total
        for i in range(N):
            v = d['norm_yr_sums'][i]
            if pd.notna(v): g_yr[i] += v
        if pd.notna(d['sum_avg10_norm']): g_avg += d['sum_avg10_norm']
        if pd.notna(d['sum_esi']):        g_esi += d['sum_esi']

    # Grand Total row
    ws.cell(row=r_total, column=1).value = 'Grand Total'
    for i in range(N): ws.cell(row=r_total, column=2+i).value = _nan_none(g_yr[i])
    ws.cell(row=r_total, column=12).value = _nan_none(g_avg)
    ws.cell(row=r_total, column=13).value = _nan_none(g_esi)
    return r_total + 3   # 2-row gap before next section


def calculate_corridor(df_cupncap: pd.DataFrame, si_col: str,
                       floor: float, cap: float) -> Workbook:
    """
    Build the corridor workbook for a given floor/cap pair.
    Each sheet has up to 3 stacked sections: Overall, Kharif, Rabi.
      - Overall : all seasons combined — no season filter
      - Kharif  : only Kharif rows — own pivot and own calc block
      - Rabi    : only Rabi rows   — own pivot and own calc block
    Sections with no data are skipped automatically.
    """
    floor_pct = int(round(floor * 100))
    cap_pct   = int(round(cap   * 100))
    wb        = Workbook(); wb.remove(wb.active)
    ws        = wb.create_sheet(f'{floor_pct}-{cap_pct} {si_col}'[:31])

    # Define the three sections: (label shown in sheet, season_filter for _agg_by_cluster)
    sections = [
        (None,      None),       # Overall — no season filter, no section label row
        ('Kharif',  'Kharif'),   # Kharif only
        ('Rabi',    'Rabi'),     # Rabi only
    ]

    next_row = 1
    for section_label, season_filter in sections:
        pivot_df = _agg_by_cluster(df_cupncap, si_col, season_filter)
        if pivot_df.empty:
            continue   # skip sections that have no data

        # Build per-cluster metrics dict for this section
        clusters_data = {
            row['Cluster']: _cluster_metrics(row, si_col, floor, cap)
            for _, row in pivot_df.iterrows()
        }

        next_row = _write_section(
            ws, next_row, clusters_data,
            floor_pct, cap_pct,
            section_label=section_label,
        )

    # Auto-fit column widths
    for col_cells in ws.columns:
        mx = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        if mx: ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(mx+2, 22)
    return wb

# ════════════════════════════════════════════════════════════════════════════
#  RUNNER
# ════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    if not os.path.isfile(INPUT_PATH):
        raise FileNotFoundError(f"Input file not found:\n  {INPUT_PATH}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    stem        = os.path.splitext(os.path.basename(INPUT_PATH))[0]
    timestamp   = datetime.datetime.now().strftime('%Y-%m-%d_%H%M%S')
    output_file = os.path.join(OUTPUT_DIR, f"{stem}_output_{timestamp}.xlsx")

    print(f"\n{'='*60}\n  Crop Insurance Pricing Calculator\n{'='*60}")
    print(f"  Input : {INPUT_PATH}\n  SI    : {SI_PATH}")

    df_in    = (pd.read_csv(INPUT_PATH) if INPUT_PATH.lower().endswith('.csv')
                else pd.read_excel(INPUT_PATH))
    detected = detect_year_columns(df_in)
    id_det   = detect_id_columns(df_in, detected)
    print(f"  Rows  : {len(df_in):,}  |  Years: {detected[0]}-{detected[-1]}  |  IDs: {id_det}")

    print("\n  [1/5] Base Data & T Test ...")
    df_base = calculate_all(df_in)
    print(f"        {len(df_base):,} rows x {len(df_base.columns)} cols")

    print("  [2/5] LossCost ...")
    df_losscost = calculate_losscost(df_base)
    print(f"        {len(df_losscost):,} rows x {len(df_losscost.columns)} cols")

    print("  [3/5] BC ...")
    df_bc = calculate_bc(df_losscost)

    print("  [4/5] Suminsured ...")
    df_si   = load_suminsured(SI_PATH)
    si_cols = [c for c in df_si.columns if c not in SI_JOIN_KEYS]

    print("  [5/5] CupnCap & Corridor sheets ...")
    cupncap_sheets: Dict[str, pd.DataFrame] = {}
    corridor_wbs:   Dict[str, Workbook]     = {}

    for si_col in si_cols:
        df_cc = calculate_cupncap(df_losscost, df_si, si_col)
        cupncap_sheets[si_col] = df_cc
        print(f"        CupnCapModel_{si_col}: {len(df_cc):,} rows")
        for floor_pct, cap_pct in CORRIDOR_MODELS:
            wb_corr    = calculate_corridor(df_cc, si_col, floor_pct/100, cap_pct/100)
            sheet_name = f'{floor_pct}-{cap_pct} {si_col[:12]}'[:31]
            corridor_wbs[sheet_name] = wb_corr
            print(f"        {sheet_name}: done")

    print("\n  Writing output ...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_base.to_excel(writer,     index=False, sheet_name='Base Data & T Test')
        df_losscost.to_excel(writer, index=False, sheet_name='LossCost')
        df_bc.to_excel(writer,       index=False, sheet_name='BC')
        df_si.to_excel(writer,       index=False, sheet_name='Suminsured')
        for si_col, df_cc in cupncap_sheets.items():
            df_cc.to_excel(writer, index=False, sheet_name=f'CupnCapModel_{si_col}'[:31])
        main_wb = writer.book
        for sheet_name, wb_corr in corridor_wbs.items():
            src_ws = wb_corr.active
            new_ws = main_wb.create_sheet(title=sheet_name)
            for row in src_ws.iter_rows():
                for cell in row:
                    new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            for col_letter, dim in src_ws.column_dimensions.items():
                new_ws.column_dimensions[col_letter].width = dim.width

    all_sheets = (['Base Data & T Test', 'LossCost', 'BC', 'Suminsured']
                + [f'CupnCapModel_{c}'[:31] for c in si_cols]
                + list(corridor_wbs.keys()))
    print(f"\n  Done!\n  Output : {output_file}")
    print(f"  Sheets : {' | '.join(all_sheets)}\n{'='*60}\n")
